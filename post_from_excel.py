"""
Facebook Auto-Poster — Cloudflare R2 (Bucket) + Excel Version
===============================================================
Excel থেকে schedule পড়ে, Cloudflare R2 bucket-এর images/ ও videos/
ফোল্ডার থেকে ফাইলের নাম দিয়ে ফাইল নামায়, Facebook Page-এ পোস্ট করে।

কীভাবে ব্যবহার করবে:
1. Cloudflare R2-তে একটা bucket বানাও (যেমন: fb-lol-cringe)
2. bucket-এর ভেতরে "images" ও "videos" নামে দুটো ফোল্ডার (prefix) বানাও
3. ছবি images/ ফোল্ডারে, ভিডিও videos/ ফোল্ডারে আপলোড করো
4. Excel-এ শুধু ফাইলের নাম লিখো (কোনো prefix ছাড়া), যেমন: swert.jpg
5. Type কলামে Image/Video ঠিকভাবে দাও — এটা দিয়েই ঠিক করা হবে কোন
   ফোল্ডারে (images/ বা videos/) খুঁজবে
6. Schedule Time দাও → GitHub-এ upload করো

প্রয়োজনীয় GitHub Secrets:
  FACEBOOK_PAGE_ID
  FACEBOOK_ACCESS_TOKEN
  R2_ACCESS_KEY_ID       ← Cloudflare R2 API token থেকে
  R2_SECRET_ACCESS_KEY   ← Cloudflare R2 API token থেকে
  R2_ENDPOINT_URL        ← https://<account_id>.r2.cloudflarestorage.com
  R2_BUCKET_NAME         ← যেমন: fb-lol-cringe

⚠️  bucket-টা public হতে হবে না — boto3 নিজের access key/secret দিয়ে
    সরাসরি R2-র সাথে কথা বলে, তাই Google Drive-এর মতো "Anyone with
    the link" শেয়ার করার দরকার নেই।
"""

import os
import sys
import requests
import openpyxl
import boto3
import tempfile
from pathlib import Path
from datetime import datetime, timezone, timedelta

# ── Config ────────────────────────────────────────────
PAGE_ID       = os.environ.get("FACEBOOK_PAGE_ID",      "YOUR_PAGE_ID")
ACCESS_TOKEN  = os.environ.get("FACEBOOK_ACCESS_TOKEN", "YOUR_TOKEN")

R2_ACCESS_KEY = os.environ.get("R2_ACCESS_KEY_ID", "")
R2_SECRET_KEY = os.environ.get("R2_SECRET_ACCESS_KEY", "")
R2_ENDPOINT   = os.environ.get("R2_ENDPOINT_URL", "")
R2_BUCKET     = os.environ.get("R2_BUCKET_NAME", "")

EXCEL_FILE  = Path("facebook_content_calendar.xlsx")
SHEET_NAME  = "Content Calendar"

COL_ID       = 1
COL_FILENAME = 2
COL_TYPE     = 3
COL_CAPTION  = 4
COL_SCHEDULE = 5
COL_STATUS   = 6
COL_POST_ID  = 7
COL_NOTE     = 8

BASE_URL = f"https://graph.facebook.com/v19.0/{PAGE_ID}"
IST      = timezone(timedelta(hours=5, minutes=30))

s3 = boto3.client(
    "s3",
    endpoint_url=R2_ENDPOINT,
    aws_access_key_id=R2_ACCESS_KEY,
    aws_secret_access_key=R2_SECRET_KEY,
)

# ── Cloudflare R2 helpers ──────────────────────────────

def download_from_r2(post_type, filename, dest_path):
    """R2 bucket-এর images/ বা videos/ ফোল্ডার থেকে ফাইল নামায়।"""
    prefix = "images" if post_type == "image" else "videos"
    key = f"{prefix}/{filename}"
    try:
        s3.download_file(R2_BUCKET, key, dest_path)
    except Exception as e:
        print(f"  ❌ '{key}' ডাউনলোড ব্যর্থ: {e}")
        return False

    size = Path(dest_path).stat().st_size
    if size == 0:
        print("  ❌ ডাউনলোড হওয়া ফাইল খালি")
        return False

    print(f"  ✓ R2 থেকে download হয়েছে ({size//1024} KB)")
    return True

# ── Load Excel ────────────────────────────────────────

def load_sheet():
    if not EXCEL_FILE.exists():
        print(f"❌ Excel file পাওয়া যায়নি: {EXCEL_FILE}")
        sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    return wb, ws

def save_sheet(wb):
    wb.save(EXCEL_FILE)

def get_due_rows(ws):
    now = datetime.now(IST).replace(tzinfo=None)
    due = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        status   = row[COL_STATUS - 1].value
        schedule = row[COL_SCHEDULE - 1].value
        if status and str(status).strip().lower() == "pending":
            if schedule and isinstance(schedule, datetime) and schedule <= now:
                due.append(row)
    return due

# ── Post functions ────────────────────────────────────

def post_image(file_path, caption):
    print(f"  📸 Image পোস্ট করছি...")
    with open(file_path, "rb") as f:
        resp = requests.post(
            f"{BASE_URL}/photos",
            params={"access_token": ACCESS_TOKEN},
            files={"source": f},
            data={"caption": caption or "", "published": "true"}
        ).json()
    if "id" in resp:
        return resp["id"]
    print(f"  ❌ Image error: {resp.get('error', {}).get('message', resp)}")
    return None

def post_video(file_path, caption):
    print(f"  🎬 Video পোস্ট করছি...")
    file_size = Path(file_path).stat().st_size

    init = requests.post(
        f"{BASE_URL}/video_reels",
        data={"upload_phase": "start", "access_token": ACCESS_TOKEN}
    ).json()
    if "error" in init:
        print(f"  ❌ Video init error: {init['error']['message']}")
        return None

    video_id   = init["video_id"]
    upload_url = init["upload_url"]

    with open(file_path, "rb") as f:
        up = requests.post(
            upload_url,
            headers={
                "Authorization": f"OAuth {ACCESS_TOKEN}",
                "offset": "0",
                "file_size": str(file_size),
            },
            data=f
        ).json()
    if not up.get("success"):
        print(f"  ❌ Upload error: {up}")
        return None

    pub = requests.post(
        f"{BASE_URL}/video_reels",
        data={
            "upload_phase": "finish",
            "access_token": ACCESS_TOKEN,
            "video_id": video_id,
            "video_state": "PUBLISHED",
            "description": caption or "",
        }
    ).json()
    if "error" in pub:
        print(f"  ❌ Publish error: {pub['error']['message']}")
        return None
    return video_id

def post_text(caption):
    print(f"  📝 Text পোস্ট করছি...")
    resp = requests.post(
        f"{BASE_URL}/feed",
        data={"message": caption or "", "access_token": ACCESS_TOKEN}
    ).json()
    if "id" in resp:
        return resp["id"]
    print(f"  ❌ Text error: {resp.get('error', {}).get('message', resp)}")
    return None

def mark_done(ws, row, post_id):
    from openpyxl.styles import PatternFill, Font
    row[COL_STATUS - 1].value  = "Done"
    row[COL_STATUS - 1].fill   = PatternFill("solid", fgColor="D4EDDA")
    row[COL_STATUS - 1].font   = Font(name="Arial", size=10, bold=True, color="155724")
    row[COL_POST_ID - 1].value = str(post_id)

# ── Main ──────────────────────────────────────────────

def main():
    now_ist = datetime.now(IST)
    print("=" * 55)
    print("  Facebook Auto-Poster (Cloudflare R2 + Excel)")
    print(f"  সময়: {now_ist.strftime('%d/%m/%Y %H:%M')} IST")
    print("=" * 55)

    wb, ws = load_sheet()
    due_rows = get_due_rows(ws)

    if not due_rows:
        print("ℹ️  এই মুহূর্তে কোনো scheduled পোস্ট নেই।")
        return

    print(f"\n📋 {len(due_rows)}টি পোস্ট পাওয়া গেছে।\n")
    posted_count = 0

    for row in due_rows:
        row_num   = row[0].row
        filename  = str(row[COL_FILENAME - 1].value or "").strip()
        post_type = str(row[COL_TYPE - 1].value or "").strip().lower()
        caption   = str(row[COL_CAPTION - 1].value or "").strip()

        print(f"── Row {row_num}: {filename or '(text only)'} [{post_type}]")

        post_id = None

        if post_type == "text" or not filename:
            post_id = post_text(caption)

        elif post_type in ("image", "video"):
            if not filename:
                print(f"  ⚠️  ফাইলের নাম খালি, স্কিপ করা হলো")
                continue

            if not R2_BUCKET:
                print(f"  ❌ R2_BUCKET_NAME সেট করা নেই")
                continue

            suffix = Path(filename).suffix or (".jpg" if post_type == "image" else ".mp4")
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                tmp_path = tmp.name

            print(f"  📥 R2 থেকে নামাচ্ছি...")
            if not download_from_r2(post_type, filename, tmp_path):
                Path(tmp_path).unlink(missing_ok=True)
                continue

            if post_type == "image":
                post_id = post_image(tmp_path, caption)
            else:
                post_id = post_video(tmp_path, caption)

            Path(tmp_path).unlink(missing_ok=True)

        else:
            print(f"  ⚠️  অচেনা ধরন: '{post_type}'")
            continue

        if post_id:
            mark_done(ws, row, post_id)
            print(f"  ✅ সফল! Post ID: {post_id}")
            posted_count += 1
        else:
            print(f"  ❌ ব্যর্থ।")

    save_sheet(wb)
    print(f"\n{'='*55}")
    print(f"  ✅ {posted_count}/{len(due_rows)} পোস্ট সফল।")
    print(f"{'='*55}")

if __name__ == "__main__":
    main()